from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from .models import Attendance, Course
from datetime import datetime
from io import BytesIO
from urllib.parse import quote
import pytz
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required

#QRコード読み込み
@login_required
def qr_reader_view(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    context = {
        'course': course,
    }
    return render(request, 'attendances/qr_reader.html', context)

#出席保存
def save_attendance(request):
    if request.method == 'POST':
        qr_data = request.POST.get('qr_data', '')
        course_id = request.POST.get('course_id', '')

        # QRコードのデータを解析
        user_info = qr_data.split(',')
        if len(user_info) == 2:
            student_username = user_info[0].replace('User Info: ', '').strip()
            student_email = user_info[1]

            try:
                course = Course.objects.get(id=course_id)
                
                # 日本時間のタイムゾーンを設定
                tokyo_tz = pytz.timezone('Asia/Tokyo')

                # 現在の日本時間を取得
                attendance_time = timezone.now().astimezone(tokyo_tz)

                # 授業開始日時の設定と日本時間のタイムゾーン付加
                course_start_datetime = datetime.combine(course.start_date, course.start_time)
                course_start_datetime = tokyo_tz.localize(course_start_datetime)

                # デバッグ情報出力
                print(f"Attendance time: {attendance_time}")
                print(f"Course start datetime: {course_start_datetime}")

                # 出席ステータスの決定
                time_diff = (attendance_time - course_start_datetime).total_seconds() / 60  # 分単位
                print(f"Time difference: {time_diff} minutes")

                if time_diff <= 10:
                    status = 'Present'
                elif time_diff <= 15:
                    status = 'Late'
                else:
                    status = 'Absent'

                # 重複チェック
                if not Attendance.objects.filter(course=course, student=student_username).exists():
                    # 出席記録の保存
                    Attendance.objects.create(
                        course=course,
                        student=student_username,
                        attendance_time=attendance_time,
                        attendance=status
                    )

                    return JsonResponse({'status': 'success'})
                else:
                    return JsonResponse({'status': 'error', 'message': '既に出席が記録されています'})
            except Course.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': '授業が見つかりません'})
        else:
            return JsonResponse({'status': 'error', 'message': 'QRコードのデータが無効です'})
    else:
        return JsonResponse({'status': 'error', 'message': 'POSTメソッドのみ対応しています'})

#出席リスト表示
@login_required
def attendance_list_view(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    attendances = Attendance.objects.filter(course=course).order_by('attendance_time')

    context = {
        'course': course,
        'attendances': attendances,
    }
    return render(request, 'attendances/attendance_list.html', context)

#出席リスト(エクセルエクスポート)
@login_required
def export_attendance_to_excel(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    attendances = Attendance.objects.filter(course=course).order_by('attendance_time')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"{course.name} 出席リスト"

    # 設定するスタイル
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                    top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    # 色の設定
    fill_present = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 緑色
    fill_late = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")     # 黄色
    fill_absent = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # 赤色

    # 授業情報の書き込み
    sheet["A1"] = "授業名:"
    sheet["B1"] = course.name
    sheet["A2"] = "授業日時:"
    sheet["B2"] = f"{course.start_date} {course.start_time.strftime('%H:%M:%S')}"
    sheet["A3"] = "教師:"
    sheet["B3"] = course.teacher.username

    # ヘッダー行の書き込みとスタイル適用
    headers = ["学籍番号", "出席時間", "出席"]
    sheet.append(headers)
    for cell in sheet[4]:
        cell.fill = header_fill
        cell.border = border

    # 出席ステータスの日本語マッピング
    status_mapping = {
        'Present': ('出席', fill_present),
        'Late': ('遅刻', fill_late),
        'Absent': ('欠席', fill_absent),
    }

    # 出席データの追加とスタイル適用
    max_lengths = [len(header) for header in headers]  # ヘッダーの長さを基に最大長を設定
    for row_idx, attendance in enumerate(attendances, start=5):
        status_japanese, fill = status_mapping.get(attendance.attendance, ('未確認', None))
        row = [
            attendance.student,
            attendance.attendance_time.strftime("%Y-%m-%d %H:%M:%S"),
            status_japanese
        ]
        sheet.append(row)
        
        # 最大列幅を更新
        for col_idx, cell in enumerate(row, start=1):
            max_lengths[col_idx - 1] = max(max_lengths[col_idx - 1], len(str(cell)))

        # 行の枠線と背景色を適用
        for col_idx, cell in enumerate(sheet[row_idx], start=1):
            cell.border = border
            if col_idx == 3:  # 出席列に色を適用
                cell.fill = fill

    # 列幅の設定
    for col_idx, length in enumerate(max_lengths, start=1):
        # '出席'列の幅を特別に広げる
        if col_idx == 3:
            sheet.column_dimensions[get_column_letter(col_idx)].width = length + 5  # 出席列に余裕を持たせる
        else:
            sheet.column_dimensions[get_column_letter(col_idx)].width = length + 2  # 他の列

    # Excelファイルをバイトストリームに保存
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)

    # ファイル名をURLエンコード
    encoded_filename = quote(f"{course.name}_attendance.xlsx")

    # レスポンスの生成
    response = HttpResponse(
        content=excel_stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
    
    return response